# =========================================================
# 냉동 CT 온도관리 플랫폼 V4.4
# - 기준일: 오늘(KST) 고정 (의사결정 사항 / 데이터 미수신 감지 목적)
# - 요구 Streamlit >= 1.40 (st.pills, st.segmented_control)
# - cold_chain_report_payload / cold_chain_webhook 인터페이스 불변
# =========================================================
import io
import re
import numpy as np
import pandas as pd
import streamlit as st
import matplotlib.pyplot as plt
import matplotlib.dates as mdates
import plotly.graph_objects as go
import cold_chain_report_payload as report_payload
import cold_chain_webhook as webhook_client
import streamlit.components.v1 as components
from openpyxl.drawing.image import Image as XLImage
from openpyxl.utils import get_column_letter

# =========================================================
# 판정 기준 상수 (기준 변경 시 이 블록만 수정)
# =========================================================
TH = {
    "abnormal_low": -50.0,     # 이상값: 미만 → 결측 처리
    "abnormal_high": 60.0,     # 이상값: 초과 → 결측 처리
    "min_temp_off": -16.0,     # 최저온도 이 값 이상이면 이탈
    "eff_low": 70.0,           # 냉동효율 미만이면 이탈
    "eff_high": 120.0,         # 냉동효율 초과면 과냉(참고지표)
    "eff_emergency": 60.0,     # 냉동효율 이하면 긴급점검
    "retention_off": 70.0,     # -15℃ 유지율 이하면 이탈
    "retention_temp": -15.0,   # 유지율 판정 온도
    "target_temp": 18.0,       # 목표면적 계수 (0-(-18)℃)
}

# =========================================================
# 상태 메타 (심각도 순 색상 위계: 긴급점검이 가장 강함)
# =========================================================
STATUS_META = {
    "정상": {
        "icon": "✓", "severity": 0,
        "bg": "#0f2a1a", "border": "#22c55e", "text": "#bbf7d0", "badge": "#16a34a",
    },
    "주의": {
        "icon": "⚠", "severity": 1,
        "bg": "#332a08", "border": "#eab308", "text": "#fde68a", "badge": "#ca8a04",
    },
    "위험": {
        "icon": "▲", "severity": 2,
        "bg": "#3a1c07", "border": "#f97316", "text": "#fed7aa", "badge": "#ea580c",
    },
    "긴급점검": {
        "icon": "✕", "severity": 3,
        "bg": "#450a0a", "border": "#ef4444", "text": "#fecaca", "badge": "#dc2626",
    },
    "데이터 연결 이상": {
        "icon": "◌", "severity": 4,
        "bg": "#1f2937", "border": "#9ca3af", "text": "#e5e7eb", "badge": "#6b7280",
    },
}

WEEKDAY_KR = {
    "Monday": "월요일", "Tuesday": "화요일", "Wednesday": "수요일",
    "Thursday": "목요일", "Friday": "금요일", "Saturday": "토요일", "Sunday": "일요일",
}

METRIC_LABEL_EN = {
    "최저온도": "Min Temperature (°C)",
    "냉동효율(%)": "Freezing Efficiency (%)",
    "-15℃이하유지율(%)": "Below -15°C Retention (%)",
}

# =========================================================
# Streamlit 기본 설정
# =========================================================
st.set_page_config(
    page_title="❄ 냉동 CT 온도관리 시스템",
    layout="wide"
)

# =========================================================
# 공통 UI 스타일
# =========================================================
st.markdown(
    """
    <style>
    html {
        scroll-behavior: smooth;
    }
    .section-anchor {
        display: block;
        position: relative;
        top: -90px;
        visibility: hidden;
    }
    .sidebar-nav-title {
        font-size: 0.86rem;
        font-weight: 800;
        margin-top: 1.1rem;
        margin-bottom: 0.35rem;
        color: #9ca3af;
    }
    .sidebar-nav a {
        display: block;
        padding: 0.28rem 0.1rem;
        color: #e5e7eb;
        text-decoration: none;
        font-size: 0.86rem;
        font-weight: 600;
        line-height: 1.35;
    }
    .sidebar-nav a:hover {
        color: #ffffff;
        text-decoration: underline;
    }
    .cc-action-card {
        border: 1px solid rgba(148, 163, 184, 0.35);
        border-radius: 14px;
        padding: 18px 18px 14px 18px;
        background: rgba(15, 23, 42, 0.55);
        min-height: 178px;
        margin-bottom: 12px;
    }
    .cc-action-card-title {
        font-size: 1.15rem;
        font-weight: 850;
        color: #ffffff;
        margin-bottom: 8px;
    }
    .cc-action-card-desc {
        font-size: 0.9rem;
        color: #cbd5e1;
        line-height: 1.55;
        margin-bottom: 12px;
    }
    .cc-mini-label {
        font-size: 0.78rem;
        color: #94a3b8;
        font-weight: 700;
        margin-bottom: 2px;
    }
    .cc-mini-value {
        font-size: 1.25rem;
        color: #ffffff;
        font-weight: 850;
    }
    .cc-step-note {
        border-radius: 12px;
        padding: 12px 14px;
        background: rgba(30, 64, 175, 0.28);
        border: 1px solid rgba(59, 130, 246, 0.35);
        color: #bfdbfe;
        font-size: 0.92rem;
        font-weight: 650;
        margin-bottom: 16px;
    }
    .cc-result-box {
        border: 1px solid rgba(148, 163, 184, 0.28);
        border-radius: 14px;
        padding: 14px 16px;
        background: rgba(15, 23, 42, 0.45);
        margin-top: 8px;
        margin-bottom: 14px;
    }
    /* 상단 요약 칩 */
    .cc-chip-row {
        display: flex;
        gap: 10px;
        flex-wrap: wrap;
        align-items: stretch;
        margin: 6px 0 14px 0;
    }
    .cc-chip {
        border-radius: 12px;
        padding: 10px 16px;
        min-width: 108px;
        border: 1px solid rgba(148, 163, 184, 0.3);
        background: rgba(15, 23, 42, 0.5);
    }
    .cc-chip-label {
        font-size: 0.78rem;
        font-weight: 700;
        color: #94a3b8;
        margin-bottom: 2px;
    }
    .cc-chip-value {
        font-size: 1.25rem;
        font-weight: 850;
        color: #e5e7eb;
    }
    .cc-chip.alert .cc-chip-value {
        font-size: 1.7rem;
    }
    </style>
    """,
    unsafe_allow_html=True,
)


def section_anchor(anchor_id: str):
    st.markdown(
        f'<span id="{anchor_id}" class="section-anchor"></span>',
        unsafe_allow_html=True,
    )


with st.sidebar:
    st.markdown("### ❄ 냉동 CT 관리")
    menu = st.radio(
        "메뉴",
        [
            "분석 프로그램",
            "자동보고서 관리",
        ],
        index=0
    )
    st.markdown("---")
    if menu == "분석 프로그램":
        st.caption("분석 결과는 상단 탭(대시보드 / 상세 분석 / 그래프)으로 이동할 수 있습니다.")
    else:
        st.markdown(
            """
            <div class="sidebar-nav-title">📌 화면 바로가기</div>
            <div class="sidebar-nav">
              <a href="#auto-summary-section">현재 분석 상태</a>
              <a href="#auto-action-section">실행 버튼</a>
              <a href="#store-result-section">Store 결과</a>
              <a href="#preview-result-section">보고서 미리보기</a>
              <a href="#send-section">최종 메일 발송</a>
              <a href="#admin-section">관리자 / 테스트 기능</a>
            </div>
            """,
            unsafe_allow_html=True,
        )

# =========================================================
# Session State 초기화
# =========================================================
SESSION_DEFAULTS = {
    "analysis_done": False,
    "report_payload_json": None,
    "report_meta": None,
    "summary_payload": None,
    "heatmap_rows_payload": None,
    "check_list_payload": None,
    "metrics_payload": None,
    "data_quality_payload": None,
    "daily_summary_rows_payload": None,
    "webhook_ping_result": None,
    "webhook_ping_ok": False,
    "draft_result": None,
    "draft_ok": False,
    "preview_result": None,
    "preview_ok": False,
    "payload_test_result": None,
    "payload_test_ok": False,
    "main_preview_result": None,
    "main_preview_ok": False,
    "send_confirmed": False,
    "send_result": None,
    "send_ok": False,
    "store_result": None,
    "store_ok": False,
    "store_attempted": False,
    "main_preview_attempted": False,
}
for _key, _value in SESSION_DEFAULTS.items():
    st.session_state.setdefault(_key, _value)


# =========================================================
# 관리자 / 테스트 패널 (분석 전후 공용)
# =========================================================
def render_admin_panel(webhook_url, webhook_token, report_payload_json=None, key_prefix="admin"):
    with st.expander("🔧 관리자 / 테스트 기능", expanded=False):
        st.caption("일반 운영자는 열지 않아도 됩니다. 연결 점검, 전송 데이터 확인용입니다.")
        st.markdown("#### 1) 시스템 연결 테스트")
        col_a, col_b = st.columns(2)
        with col_a:
            if webhook_url:
                st.success("Webhook URL 설정됨")
            else:
                st.error("Webhook URL 미설정")
        with col_b:
            if webhook_token:
                st.success("Webhook Token 설정됨")
            else:
                st.error("Webhook Token 미설정")
        if st.button("시스템 연결 테스트", key=f"{key_prefix}_ping_button", use_container_width=True):
            ping_result = webhook_client.post_webhook_ping(
                webhook_url=webhook_url,
                webhook_token=webhook_token,
            )
            st.session_state["webhook_ping_result"] = ping_result
            st.session_state["webhook_ping_ok"] = bool(ping_result.get("ok"))
        ping_result = st.session_state.get("webhook_ping_result")
        if ping_result is not None:
            if st.session_state.get("webhook_ping_ok"):
                st.success("✅ Apps Script Webhook 연결 성공")
            else:
                st.error("❌ Apps Script Webhook 연결 실패")
            if st.checkbox("연결 테스트 응답 JSON 보기", key=f"{key_prefix}_show_ping_json"):
                st.json(ping_result)

        if report_payload_json is None:
            return

        st.markdown("---")
        st.markdown("#### 2) 데이터 전송 검증")
        st.caption("현재 payload가 Apps Script에서 정상 수신되는지 확인합니다.")
        if st.button("데이터 전송 검증", key=f"{key_prefix}_payload_test_button", use_container_width=True):
            with st.spinner("Payload 수신 상태를 확인 중입니다..."):
                payload_test_result = webhook_client.post_report_preview(
                    webhook_url=webhook_url,
                    webhook_token=webhook_token,
                    payload=report_payload_json,
                    timeout=60,
                )
            st.session_state["payload_test_result"] = payload_test_result
            st.session_state["payload_test_ok"] = bool(payload_test_result.get("ok"))
        payload_test_result = st.session_state.get("payload_test_result")
        if payload_test_result is not None:
            if st.session_state.get("payload_test_ok"):
                st.success("✅ 데이터 전송 검증 성공")
                response = payload_test_result.get("response", {})
                received = response.get("received", {})
                col_p1, col_p2, col_p3, col_p4 = st.columns(4)
                with col_p1:
                    st.metric("schema", received.get("schema_version", "-"))
                with col_p2:
                    st.metric("heatmap_rows", received.get("heatmap_rows_count", 0))
                with col_p3:
                    st.metric("check_list", received.get("check_list_count", 0))
                with col_p4:
                    st.metric("metrics", received.get("metrics_count", 0))
                st.caption(
                    f"mode: {received.get('mode', '-')} / "
                    f"report_date: {received.get('report_date', '-')}"
                )
            else:
                st.error("❌ 데이터 전송 검증 실패")
            if st.checkbox("데이터 전송 검증 응답 JSON 보기", key=f"{key_prefix}_show_payload_test_json"):
                st.json(payload_test_result)

        st.markdown("---")
        st.markdown("#### 3) 전송 데이터 상세 보기")
        if st.checkbox("저장된 report payload JSON 보기", key=f"{key_prefix}_show_saved_payload_json"):
            st.json(report_payload_json)


# =========================================================
# 자동보고서 관리 페이지
# =========================================================
def render_auto_report_page():
    st.title("📧 냉동 CT 자동보고서 관리")
    st.caption("분석 프로그램에서 CSV 분석을 완료한 뒤 Store, 미리보기, 최종 발송을 진행합니다.")
    webhook_url = st.secrets.get("APPS_SCRIPT_WEBHOOK_URL", "")
    webhook_token = st.secrets.get("REPORT_WEBHOOK_TOKEN", "")

    # 분석 전 안내 화면
    if not st.session_state.get("analysis_done"):
        st.warning("먼저 좌측 메뉴의 [분석 프로그램]에서 CSV 업로드 및 분석을 완료하세요.")
        render_admin_panel(webhook_url, webhook_token, key_prefix="admin_pre")
        return

    # 분석 완료 후 payload 불러오기
    report_meta = st.session_state.get("report_meta") or {}
    summary_payload = st.session_state.get("summary_payload") or {}
    check_list_payload = st.session_state.get("check_list_payload") or []
    heatmap_rows_payload = st.session_state.get("heatmap_rows_payload") or []
    metrics_payload = st.session_state.get("metrics_payload") or []
    data_quality_payload = st.session_state.get("data_quality_payload") or {}
    daily_summary_rows_payload = st.session_state.get("daily_summary_rows_payload") or []
    report_payload_json = st.session_state.get("report_payload_json")

    webhook_ready = bool(webhook_url) and bool(webhook_token)
    payload_ready = report_payload_json is not None
    daily_rows_ready = len(daily_summary_rows_payload) > 0
    can_store_dataset = webhook_ready and payload_ready and daily_rows_ready
    can_generate_report = webhook_ready and payload_ready

    # -------------------------------------------------
    # 현재 분석 상태
    # -------------------------------------------------
    section_anchor("auto-summary-section")
    st.markdown("### 현재 분석 상태")
    st.markdown(
        """
        <div class="cc-step-note">
        권장 순서: ① Store / 공식 데이터셋 적재 → ② 자동보고서 생성 / 미리보기 → ③ 최종 메일 발송
        </div>
        """,
        unsafe_allow_html=True,
    )
    col1, col2, col3, col4, col5 = st.columns(5)
    with col1:
        st.metric("기준일", report_meta.get("report_date", "-"))
    with col2:
        st.metric("분석대상", f"{summary_payload.get('total_ct', 0)}대")
    with col3:
        st.metric("우선점검", f"{len(check_list_payload)}대")
    with col4:
        st.metric("daily rows", f"{len(daily_summary_rows_payload):,}건")
    with col5:
        st.metric("유효 데이터", f"{data_quality_payload.get('valid_count', 0):,}건")
    st.caption(
        f"분석기간: {report_meta.get('period_start', '-')} ~ {report_meta.get('period_end', '-')} / "
        f"히트맵 {len(heatmap_rows_payload)}건 / "
        f"metrics {len(metrics_payload)}건 / "
        f"업로드 파일: {report_meta.get('uploaded_filename', '-')}"
    )
    if check_list_payload:
        check_cts = [item.get("ct_label", "") for item in check_list_payload]
        st.write("우선점검 CT: " + ", ".join(check_cts))
    else:
        st.write("우선점검 CT: 해당 없음")
    st.divider()

    # -------------------------------------------------
    # 핵심 실행 버튼 2열
    # -------------------------------------------------
    section_anchor("auto-action-section")
    st.markdown("### 실행")
    action_col_1, action_col_2 = st.columns(2, gap="large")
    with action_col_1:
        st.markdown(
            f"""
            <div class="cc-action-card">
                <div class="cc-action-card-title">💾 Store / 공식 데이터셋 적재</div>
                <div class="cc-action-card-desc">
                    현재 분석 결과의 daily_summary_rows를 Google Sheet <b>데이터 다운로드</b> 시트에 반영합니다.
                    중복 기준은 <b>컨테이너 + 측정일자</b>입니다.
                </div>
                <div class="cc-mini-label">적재 대상</div>
                <div class="cc-mini-value">{len(daily_summary_rows_payload):,}건</div>
                <div class="cc-mini-label" style="margin-top:8px;">대상 시트</div>
                <div class="cc-mini-value">데이터 다운로드</div>
            </div>
            """,
            unsafe_allow_html=True,
        )
        if st.button(
            "💾 Store 실행",
            key="store_dataset_button",
            type="secondary",
            use_container_width=True,
            disabled=not can_store_dataset,
        ):
            st.session_state["store_attempted"] = True
            st.session_state["store_result"] = None
            st.session_state["store_ok"] = False
            with st.spinner("공식 데이터셋에 적재 중입니다..."):
                store_result = webhook_client.post_report_store(
                    webhook_url=webhook_url,
                    webhook_token=webhook_token,
                    payload=report_payload_json,
                    timeout=90,
                )
            st.session_state["store_result"] = store_result
            st.session_state["store_ok"] = bool(store_result.get("ok"))
    with action_col_2:
        st.markdown(
            f"""
            <div class="cc-action-card">
                <div class="cc-action-card-title">🧾 자동보고서 생성 / 미리보기</div>
                <div class="cc-action-card-desc">
                    현재 분석 결과를 기준으로 발송 전 보고서를 생성합니다.
                    보고서 내용과 수신자를 확인한 뒤 최종 메일 발송 단계로 진행합니다.
                </div>
                <div class="cc-mini-label">Payload 상태</div>
                <div class="cc-mini-value">{"준비 완료" if payload_ready else "미준비"}</div>
                <div class="cc-mini-label" style="margin-top:8px;">Webhook 상태</div>
                <div class="cc-mini-value">{"연결 정보 있음" if webhook_ready else "설정 필요"}</div>
            </div>
            """,
            unsafe_allow_html=True,
        )
        if st.button(
            "🧾 자동보고서 생성 / 미리보기",
            key="main_preview_button",
            type="primary",
            use_container_width=True,
            disabled=not can_generate_report,
        ):
            st.session_state["main_preview_attempted"] = True
            st.session_state["main_preview_result"] = None
            st.session_state["main_preview_ok"] = False
            st.session_state["send_confirmed"] = False
            st.session_state["send_result"] = None
            st.session_state["send_ok"] = False
            with st.spinner("자동보고서를 생성 중입니다..."):
                result = webhook_client.post_report_preview(
                    webhook_url=webhook_url,
                    webhook_token=webhook_token,
                    payload=report_payload_json,
                    timeout=60,
                )
            st.session_state["main_preview_result"] = result
            st.session_state["main_preview_ok"] = bool(result.get("ok"))

    if not webhook_ready:
        st.warning("Webhook URL 또는 Token 설정이 없어 Store/자동보고서 기능을 실행할 수 없습니다.")
    if not payload_ready:
        st.warning("보고서 payload가 준비되지 않았습니다. 분석 프로그램에서 CSV 분석을 다시 완료하세요.")
    if not daily_rows_ready:
        st.warning("daily_summary_rows가 비어 있어 Store를 실행할 수 없습니다.")
    st.divider()

    # -------------------------------------------------
    # 실행 결과 영역
    # -------------------------------------------------
    st.markdown("### 실행 결과")
    store_result = st.session_state.get("store_result")
    store_attempted = st.session_state.get("store_attempted", False)
    store_ok = st.session_state.get("store_ok", False)
    main_preview_result = st.session_state.get("main_preview_result")
    main_preview_attempted = st.session_state.get("main_preview_attempted", False)
    main_preview_ok = st.session_state.get("main_preview_ok", False)

    # Store 결과
    section_anchor("store-result-section")
    st.markdown("#### 1) Store 결과")
    if store_attempted and store_result is not None:
        if store_ok:
            response = store_result.get("response", {})
            store_info = response.get("store", {})
            st.success("✅ Store 완료: 공식 데이터셋에 반영되었습니다.")
            col_store_r1, col_store_r2, col_store_r3, col_store_r4 = st.columns(4)
            with col_store_r1:
                st.metric("insert", store_info.get("insert_count", 0))
            with col_store_r2:
                st.metric("update", store_info.get("update_count", 0))
            with col_store_r3:
                st.metric("skip", store_info.get("skip_count", 0))
            with col_store_r4:
                st.metric("valid rows", store_info.get("valid_rows", 0))
            st.caption(
                f"sheet: {store_info.get('sheet_name', '데이터 다운로드')} / "
                f"run_id: {store_info.get('ingest_run_id', '-')}"
            )
        else:
            st.error("❌ Store 실패")
            error_message = store_result.get("error") or store_result.get("response", {}).get("error")
            if error_message:
                st.code(str(error_message))
        if st.checkbox("Store 응답 JSON 보기", key="show_store_result_json"):
            st.json(store_result)
    else:
        st.info("아직 Store를 실행하지 않았습니다.")
    st.divider()

    # 미리보기 결과
    section_anchor("preview-result-section")
    st.markdown("#### 2) 자동보고서 미리보기 결과")
    if main_preview_attempted:
        if main_preview_ok and main_preview_result is not None:
            st.success("✅ 자동보고서 생성 완료. 아래 내용을 확인하세요.")
            preview = main_preview_result.get("response", {}).get("preview", {})
            received = main_preview_result.get("response", {}).get("received", {})
            html_body = preview.get("html_body", "")
            col_m1, col_m2 = st.columns([1, 2])
            with col_m1:
                st.caption("수신자")
                st.write(preview.get("to", "-"))
            with col_m2:
                st.caption("제목")
                st.write(preview.get("subject", "-"))
            with st.expander("수신 데이터 요약", expanded=False):
                col_r1, col_r2, col_r3, col_r4 = st.columns(4)
                with col_r1:
                    st.metric("schema", received.get("schema_version", "-"))
                with col_r2:
                    st.metric("heatmap_rows", received.get("heatmap_rows_count", 0))
                with col_r3:
                    st.metric("check_list", received.get("check_list_count", 0))
                with col_r4:
                    st.metric("metrics", received.get("metrics_count", 0))
            if html_body:
                st.markdown("##### 보고서 미리보기")
                components.html(
                    html_body,
                    height=980,
                    scrolling=True,
                )
            else:
                st.warning("미리보기 HTML이 비어 있습니다.")
            if st.checkbox("자동보고서 생성 응답 JSON 보기", key="show_main_preview_json"):
                st.json(main_preview_result)
        else:
            st.error("❌ 자동보고서 생성 실패")
            if st.checkbox("실패 응답 JSON 보기", key="show_main_preview_error_json"):
                st.json(main_preview_result)
    else:
        st.info("아직 자동보고서 미리보기를 생성하지 않았습니다.")
    st.divider()

    # 최종 발송
    section_anchor("send-section")
    st.markdown("#### 3) 최종 메일 발송")
    if main_preview_ok and main_preview_result is not None:
        send_confirmed = st.checkbox(
            "보고서 내용과 수신자를 확인했습니다.",
            key="send_confirmed",
        )
        if send_confirmed:
            st.warning(
                "수신자와 보고서 내용을 확인했습니다. 아래 버튼을 누르면 실제 메일이 발송됩니다."
            )
        can_send_report = (
            bool(send_confirmed)
            and main_preview_ok
            and webhook_ready
            and payload_ready
            and not st.session_state.get("send_ok", False)
        )
        if st.button(
            "📨 전체 메일 발송",
            key="main_send_button",
            type="primary",
            use_container_width=True,
            disabled=not can_send_report,
        ):
            st.session_state["send_result"] = None
            st.session_state["send_ok"] = False
            with st.spinner("전체 메일을 발송 중입니다..."):
                send_result = webhook_client.post_report_send(
                    webhook_url=webhook_url,
                    webhook_token=webhook_token,
                    payload=report_payload_json,
                    timeout=90,
                )
            st.session_state["send_result"] = send_result
            st.session_state["send_ok"] = bool(send_result.get("ok"))
        send_result = st.session_state.get("send_result")
        if send_result is not None:
            if st.session_state.get("send_ok"):
                response = send_result.get("response", {})
                send_info = response.get("send", {})
                st.success("✅ 전체 메일 발송 완료")
                st.info("이미 발송 완료된 보고서입니다. 재발송하려면 CSV를 다시 분석하거나 미리보기를 새로 생성하세요.")
                col_s1, col_s2 = st.columns([1, 2])
                with col_s1:
                    st.caption("수신자")
                    st.write(send_info.get("to", "-"))
                with col_s2:
                    st.caption("제목")
                    st.write(send_info.get("subject", "-"))
                st.caption(
                    f"수신처 모드: {send_info.get('recipient_mode', '-')} / "
                    f"발송시각: {send_info.get('sent_time', '-')}"
                )
            else:
                st.error("❌ 전체 메일 발송 실패")
            if st.checkbox("메일 발송 응답 JSON 보기", key="show_send_result_json"):
                st.json(send_result)
    else:
        st.info("최종 발송은 자동보고서 미리보기 생성 후 활성화됩니다.")
    st.divider()

    # 관리자 / 테스트 기능
    section_anchor("admin-section")
    render_admin_panel(
        webhook_url,
        webhook_token,
        report_payload_json=report_payload_json,
        key_prefix="admin_post",
    )


if menu != "분석 프로그램":
    render_auto_report_page()
    st.stop()

st.title("❄ 냉동 컨테이너 온도관리 플랫폼.V4.4")
st.caption(
    f"신규 데이터로거 플랫폼 CSV 전용 | 이상값 기준: "
    f"{TH['abnormal_low']:.0f}℃ 미만 또는 {TH['abnormal_high']:.0f}℃ 초과 → 결측치 처리"
)


# =========================================================
# 공통 유틸 함수
# =========================================================
def extract_ct_number(col_name: str):
    match = re.search(r"(\d+)\s*번\s*냉동CT", str(col_name))
    return int(match.group(1)) if match else None


def ct_sort_key(ct_name: str) -> int:
    match = re.search(r"CT(\d+)", str(ct_name))
    return int(match.group(1)) if match else 9999


def integrate_trapezoid(y, x):
    if len(x) == 0:
        return 0
    if hasattr(np, "trapezoid"):
        return np.trapezoid(np.asarray(y), np.asarray(x))
    return np.trapz(np.asarray(y), np.asarray(x))


def auto_adjust_excel_column_width(worksheet):
    for col_cells in worksheet.iter_cols(min_row=1, max_row=worksheet.max_row):
        max_length = max(
            (len(str(cell.value)) for cell in col_cells if cell.value is not None),
            default=0
        )
        col_letter = get_column_letter(col_cells[0].column)
        worksheet.column_dimensions[col_letter].width = max_length + 6


def get_today_kst():
    return pd.Timestamp.now(tz="Asia/Seoul").date()


def parse_measurement_datetime(series):
    """
    데이터로거 CSV 측정일시 호환 파서.
    기존 형식과 변경된 날짜 형식을 모두 처리한다.
    """
    s = (
        series.astype(str)
        .str.strip()
        .str.replace("﻿", "", regex=False)
        .str.replace('"', "", regex=False)
    )
    parsed = pd.to_datetime(s, errors="coerce")
    if parsed.isna().any():
        try:
            parsed_mixed = pd.to_datetime(s, errors="coerce", format="mixed")
            parsed = parsed.fillna(parsed_mixed)
        except TypeError:
            pass
    if parsed.isna().any():
        s2 = (
            s.str.replace("/", "-", regex=False)
             .str.replace(".", "-", regex=False)
        )
        parsed2 = pd.to_datetime(s2, errors="coerce")
        parsed = parsed.fillna(parsed2)
    return parsed


def read_csv_flexible(file_bytes: bytes) -> pd.DataFrame:
    """
    구분자(; ,) / 인코딩(utf-8-sig, cp949) 조합 fallback 파서.
    냉동CT 컬럼이 인식되는 첫 조합을 채택한다.
    """
    attempts = [
        (";", "utf-8-sig"),
        (";", "cp949"),
        (",", "utf-8-sig"),
        (",", "cp949"),
    ]
    last_error = None
    for sep, enc in attempts:
        try:
            df = pd.read_csv(io.BytesIO(file_bytes), sep=sep, encoding=enc)
        except Exception as e:
            last_error = e
            continue
        if df.shape[1] >= 2 and any(
            extract_ct_number(col) is not None for col in df.columns[1:]
        ):
            return df
    if last_error is not None:
        raise ValueError(f"CSV 파싱 실패 (구분자/인코딩 확인 필요): {last_error}")
    raise ValueError("냉동CT 온도 컬럼을 찾지 못했습니다. CSV 컬럼명을 확인해 주세요.")


# =========================================================
# 신규 플랫폼 CSV 전처리 (캐싱: 파일 bytes 기준)
# =========================================================
@st.cache_data(show_spinner=False)
def preprocess_new_platform_csv(file_bytes: bytes):
    df_raw = read_csv_flexible(file_bytes)
    if df_raw.empty:
        raise ValueError("CSV 파일에 데이터가 없습니다.")
    timestamp_col = df_raw.columns[0]
    ct_rename = {}
    for col in df_raw.columns[1:]:
        ct_no = extract_ct_number(col)
        if ct_no is not None:
            ct_rename[col] = f"CT{ct_no}"
    if not ct_rename:
        raise ValueError("냉동CT 온도 컬럼을 찾지 못했습니다. CSV 컬럼명을 확인해 주세요.")
    df = df_raw.rename(columns={timestamp_col: "측정일시", **ct_rename})
    df["측정일시"] = parse_measurement_datetime(df["측정일시"])
    df = df.dropna(subset=["측정일시"]).copy()
    ct_cols = sorted(ct_rename.values(), key=ct_sort_key)
    df = df[["측정일시"] + ct_cols].copy()
    df_long = pd.melt(
        df,
        id_vars=["측정일시"],
        value_vars=ct_cols,
        var_name="컨테이너",
        value_name="온도"
    )
    df_long["온도"] = pd.to_numeric(df_long["온도"], errors="coerce")
    abnormal_mask = df_long["온도"].lt(TH["abnormal_low"]) | df_long["온도"].gt(TH["abnormal_high"])
    df_abnormal = (
        df_long.loc[abnormal_mask, ["측정일시", "컨테이너", "온도"]]
        .copy()
        .sort_values(["측정일시", "컨테이너"])
        .reset_index(drop=True)
    )
    abnormal_count = int(abnormal_mask.sum())
    df_long.loc[abnormal_mask, "온도"] = np.nan
    excluded_count = int(df_long["온도"].isna().sum())
    df_long = df_long.dropna(subset=["온도"]).copy()
    df_long["측정일자"] = df_long["측정일시"].dt.date
    df_long["시각"] = df_long["측정일시"].dt.strftime("%H:%M:%S")
    df_long["요일"] = df_long["측정일시"].dt.day_name()
    df_long["시간(분)"] = (
        df_long["측정일시"].dt.hour * 60
        + df_long["측정일시"].dt.minute
    )
    df_long = (
        df_long
        .sort_values(["컨테이너", "측정일시"])
        .reset_index(drop=True)
    )
    return df_raw, df_long, df_abnormal, abnormal_count, excluded_count


# =========================================================
# 요약 계산
# =========================================================
@st.cache_data(show_spinner=False)
def calculate_summary(df_long):
    summary_list = []
    for (container, date), group in df_long.groupby(["컨테이너", "측정일자"], sort=False):
        group = group.dropna(subset=["온도"]).sort_values("측정일시")
        if group.empty:
            continue
        최저온도 = group["온도"].min()
        평균온도 = group["온도"].mean()
        전체시간 = group["시간(분)"].max() - group["시간(분)"].min()
        mask = group["온도"] < 0
        tmp = group.loc[mask, ["시간(분)", "온도"]].sort_values("시간(분)")
        x = tmp["시간(분)"]
        y = 0 - tmp["온도"]
        측정면적 = integrate_trapezoid(y, x) if not x.empty else 0
        목표면적 = TH["target_temp"] * 전체시간 if 전체시간 > 0 else 0
        냉동효율 = 측정면적 / 목표면적 if 목표면적 > 0 else 0
        영하15유지율 = (group["온도"] <= TH["retention_temp"]).mean() if len(group) > 0 else np.nan
        summary_list.append({
            "컨테이너": container,
            "측정일자": date,
            "요일": group["요일"].iloc[0],
            "최저온도": round(최저온도, 1),
            "평균누적온도": round(평균온도, 1),
            "측정면적": round(측정면적, 0),
            "목표면적": round(목표면적, 0),
            "냉동효율(%)": round(냉동효율 * 100, 1),
            "-15℃이하유지율(%)": round(영하15유지율 * 100, 1),
            "측정건수": len(group)
        })
    df_summary = pd.DataFrame(summary_list)
    if df_summary.empty:
        return df_summary
    df_summary["요일"] = df_summary["요일"].map(WEEKDAY_KR)
    df_summary["측정면적"] = df_summary["측정면적"].round(0).astype(int)
    df_summary["목표면적"] = df_summary["목표면적"].round(0).astype(int)
    df_summary["측정건수"] = df_summary["측정건수"].round(0).astype(int)
    df_summary["최저온도"] = df_summary["최저온도"].round(1)
    df_summary["평균누적온도"] = df_summary["평균누적온도"].round(1)
    df_summary["냉동효율(%)"] = df_summary["냉동효율(%)"].round(1)
    df_summary["-15℃이하유지율(%)"] = df_summary["-15℃이하유지율(%)"].round(1)
    ct_order = sorted(df_summary["컨테이너"].unique(), key=ct_sort_key)
    df_summary["컨테이너"] = pd.Categorical(
        df_summary["컨테이너"],
        categories=ct_order,
        ordered=True
    )
    df_summary = (
        df_summary
        .sort_values(["컨테이너", "측정일자"])
        .reset_index(drop=True)
    )
    df_summary["컨테이너"] = df_summary["컨테이너"].astype(str)
    return df_summary


# =========================================================
# V4 상태 판정 로직
# =========================================================
def get_metric_flags(row):
    min_temp = row["최저온도"]
    eff = row["냉동효율(%)"]
    retention = row["-15℃이하유지율(%)"]
    # 품질팀 피드백 반영:
    # 과냉(효율 120% 초과)은 이탈로 보지 않고 참고지표로만 관리한다.
    temp_off = pd.notna(min_temp) and min_temp >= TH["min_temp_off"]
    eff_low = pd.notna(eff) and eff < TH["eff_low"]
    eff_high = pd.notna(eff) and eff > TH["eff_high"]
    # 이탈 판정에는 효율 저하만 반영
    eff_off = eff_low
    retention_off = pd.notna(retention) and retention <= TH["retention_off"]
    eff_emergency = pd.notna(eff) and eff <= TH["eff_emergency"]
    off_count = int(temp_off) + int(eff_off) + int(retention_off)
    issues = []
    if temp_off:
        issues.append(f"최저온도 {min_temp:.1f}℃")
    if eff_low:
        issues.append(f"냉동효율 저하 {eff:.1f}%")
    if retention_off:
        issues.append(f"-15℃ 유지율 저하 {retention:.1f}%")
    if not issues:
        issues.append("정상")
    return {
        "temp_off": temp_off,
        "eff_off": eff_off,
        "eff_low": eff_low,
        "eff_high": eff_high,  # 참고지표(과냉)
        "retention_off": retention_off,
        "eff_emergency": eff_emergency,
        "off_count": off_count,
        "issues": " / ".join(issues)
    }


def evaluate_daily_status(row):
    flags = get_metric_flags(row)
    if flags["off_count"] == 0:
        status = "정상"
        score = 0
    elif flags["eff_emergency"] or flags["off_count"] >= 2:
        status = "긴급점검"
        score = 3
    else:
        status = "주의"
        score = 1
    return status, score, flags["issues"], flags["off_count"], flags["eff_emergency"]


def calculate_deviation_score(row):
    score = 0
    min_temp = row.get("최저온도")
    eff = row.get("냉동효율(%)")
    retention = row.get("-15℃이하유지율(%)")
    if pd.notna(min_temp) and min_temp >= TH["min_temp_off"]:
        score = max(score, min_temp - TH["min_temp_off"])
    if pd.notna(eff) and eff < TH["eff_low"]:
        score = max(score, TH["eff_low"] - eff)
    if pd.notna(retention) and retention <= TH["retention_off"]:
        score = max(score, TH["retention_off"] - retention)
    return round(float(score), 1)


@st.cache_data(show_spinner=False)
def build_status_tables(df_summary, today):
    daily_status_all = df_summary.copy()
    daily_status_all["측정일자_dt"] = pd.to_datetime(daily_status_all["측정일자"])
    status_results = daily_status_all.apply(evaluate_daily_status, axis=1)
    daily_status_all["상태"] = [result[0] for result in status_results]
    daily_status_all["상태점수"] = [result[1] for result in status_results]
    daily_status_all["이슈"] = [result[2] for result in status_results]
    daily_status_all["이탈수"] = [result[3] for result in status_results]
    daily_status_all["효율긴급"] = [result[4] for result in status_results]
    daily_status_all["이탈정도"] = daily_status_all.apply(calculate_deviation_score, axis=1)
    today_ts = pd.to_datetime(today)
    start_ts = today_ts - pd.Timedelta(days=2)
    recent_df = daily_status_all[
        (daily_status_all["측정일자_dt"] >= start_ts)
        & (daily_status_all["측정일자_dt"] <= today_ts)
    ].copy()
    all_containers = sorted(df_summary["컨테이너"].unique(), key=ct_sort_key)
    container_rows = []
    check_rows = []
    for ct in all_containers:
        ct_all = daily_status_all[daily_status_all["컨테이너"] == ct].copy()
        ct_recent = recent_df[recent_df["컨테이너"] == ct].copy()
        today_row = ct_recent[ct_recent["측정일자_dt"] == today_ts].copy()
        overcool_ref = (
            bool((ct_recent["냉동효율(%)"] > TH["eff_high"]).any())
            if not ct_recent.empty else False
        )
        if today_row.empty:
            last_date = pd.to_datetime(ct_all["측정일자_dt"].max()).strftime("%Y-%m-%d") if not ct_all.empty else "-"
            status = "데이터 연결 이상"
            status_score = 4
            대표이슈 = f"오늘 데이터 없음 / 최근 측정일 {last_date}"
            today_off_count = np.nan
            min_eff = np.nan
            max_eff = np.nan
            min_retention = np.nan
            worst_min_temp = np.nan
            issue_days = int((ct_recent["이탈수"] > 0).sum()) if not ct_recent.empty else 0
            container_rows.append({
                "컨테이너": ct,
                "종합상태": status,
                "상태점수": status_score,
                "대표일자": str(today),
                "대표이슈": 대표이슈,
                "오늘이탈수": today_off_count,
                "최근3일이슈일": issue_days,
                "최악최저온도": worst_min_temp,
                "최저냉동효율": min_eff,
                "최고냉동효율": max_eff,
                "최저유지율": min_retention,
                "최근측정일": last_date,
                "과냉참고": overcool_ref,
            })
            check_rows.append({
                "상태": status,
                "컨테이너": ct,
                "측정일자": str(today),
                "이슈": 대표이슈,
                "최저온도": np.nan,
                "냉동효율(%)": np.nan,
                "-15℃이하유지율(%)": np.nan,
                "우선순위": status_score,
                "이탈정도": 999
            })
            continue
        today_record = today_row.iloc[0]
        today_flags = get_metric_flags(today_record)
        today_off_count = today_flags["off_count"]
        recent_issue_days = int((ct_recent["이탈수"] > 0).sum())
        recent_total_off = int(ct_recent["이탈수"].sum())
        recent_any_all3 = bool((ct_recent["이탈수"] == 3).any())
        recent_two_days_start = today_ts - pd.Timedelta(days=1)
        recent_two = ct_recent[ct_recent["측정일자_dt"] >= recent_two_days_start]
        recent_two_ok = (
            len(recent_two) >= 2
            and int(recent_two["이탈수"].sum()) == 0
        )
        if today_flags["eff_emergency"] or today_off_count >= 2:
            status = "긴급점검"
            status_score = 3
            대표이슈 = today_flags["issues"]
        elif recent_any_all3:
            status = "위험"
            status_score = 2
            risk_row = ct_recent[ct_recent["이탈수"] == 3].sort_values("측정일자_dt", ascending=False).iloc[0]
            대표이슈 = risk_row["이슈"]
        elif recent_total_off == 0:
            status = "정상"
            status_score = 0
            대표이슈 = "정상"
        elif recent_total_off == 1:
            status = "정상"
            status_score = 0
            issue_row = ct_recent[ct_recent["이탈수"] > 0].iloc[0]
            대표이슈 = f"단일 이탈 이력: {issue_row['이슈']}"
        elif recent_two_ok:
            status = "주의"
            status_score = 1
            대표이슈 = "최근 2일 정상 / 이전 이탈 이력 있음"
        else:
            status = "주의"
            status_score = 1
            issue_row = ct_recent.sort_values(["상태점수", "이탈정도"], ascending=[False, False]).iloc[0]
            대표이슈 = issue_row["이슈"]
        if ct_recent.empty:
            min_eff = np.nan
            max_eff = np.nan
            min_retention = np.nan
            worst_min_temp = np.nan
        else:
            min_eff = ct_recent["냉동효율(%)"].min()
            max_eff = ct_recent["냉동효율(%)"].max()
            min_retention = ct_recent["-15℃이하유지율(%)"].min()
            worst_min_temp = ct_recent["최저온도"].max()
        container_rows.append({
            "컨테이너": ct,
            "종합상태": status,
            "상태점수": status_score,
            "대표일자": str(today),
            "대표이슈": 대표이슈,
            "오늘이탈수": today_off_count,
            "최근3일이슈일": recent_issue_days,
            "최악최저온도": round(worst_min_temp, 1) if pd.notna(worst_min_temp) else np.nan,
            "최저냉동효율": round(min_eff, 1) if pd.notna(min_eff) else np.nan,
            "최고냉동효율": round(max_eff, 1) if pd.notna(max_eff) else np.nan,
            "최저유지율": round(min_retention, 1) if pd.notna(min_retention) else np.nan,
            "최근측정일": pd.to_datetime(ct_all["측정일자_dt"].max()).strftime("%Y-%m-%d") if not ct_all.empty else "-",
            "과냉참고": overcool_ref,
        })
        if status != "정상":
            check_source = today_record if status == "긴급점검" else ct_recent.sort_values(["상태점수", "이탈정도"], ascending=[False, False]).iloc[0]
            check_rows.append({
                "상태": status,
                "컨테이너": ct,
                "측정일자": pd.to_datetime(check_source["측정일자"]).strftime("%Y-%m-%d"),
                "이슈": 대표이슈,
                "최저온도": check_source["최저온도"],
                "냉동효율(%)": check_source["냉동효율(%)"],
                "-15℃이하유지율(%)": check_source["-15℃이하유지율(%)"],
                "우선순위": status_score,
                "이탈정도": check_source["이탈정도"] if "이탈정도" in check_source else calculate_deviation_score(check_source)
            })
    container_status_df = pd.DataFrame(container_rows)
    container_status_df["컨테이너정렬키"] = container_status_df["컨테이너"].apply(ct_sort_key)
    container_status_df = (
        container_status_df
        .sort_values(["상태점수", "컨테이너정렬키"], ascending=[False, True])
        .drop(columns=["컨테이너정렬키"])
        .reset_index(drop=True)
    )
    check_list_df = pd.DataFrame(check_rows)
    if not check_list_df.empty:
        check_list_df = (
            check_list_df
            .sort_values(["우선순위", "이탈정도"], ascending=[False, False])
            .drop(columns=["우선순위", "이탈정도"])
            .reset_index(drop=True)
        )
    return daily_status_all, recent_df, container_status_df, check_list_df, start_ts.date(), today


# =========================================================
# V4 대시보드 렌더링
# =========================================================
def status_style(status):
    return STATUS_META.get(status, STATUS_META["정상"])


def render_summary_chips(container_status_df, check_list_df, abnormal_count, excluded_count):
    counts = container_status_df["종합상태"].value_counts()
    total_ct = len(container_status_df)
    emergency_count = int(counts.get("긴급점검", 0))
    danger_count = int(counts.get("위험", 0))
    caution_count = int(counts.get("주의", 0))
    normal_count = int(counts.get("정상", 0))
    connection_count = int(counts.get("데이터 연결 이상", 0))

    def chip(label, value, color=None, alert=False):
        cls = "cc-chip alert" if alert else "cc-chip"
        border = f"border-color:{color};" if color else ""
        value_color = f"color:{color};" if color else ""
        return (
            f'<div class="{cls}" style="{border}">'
            f'<div class="cc-chip-label">{label}</div>'
            f'<div class="cc-chip-value" style="{value_color}">{value}</div>'
            f"</div>"
        )

    # 문제 상태를 앞에, 크게 배치 (정보 위계)
    chips_html = (
        '<div class="cc-chip-row">'
        + chip("✕ 긴급점검", f"{emergency_count}개",
               STATUS_META["긴급점검"]["border"], alert=emergency_count > 0)
        + chip("▲ 위험", f"{danger_count}개",
               STATUS_META["위험"]["border"], alert=danger_count > 0)
        + chip("◌ 연결 이상", f"{connection_count}개",
               STATUS_META["데이터 연결 이상"]["border"], alert=connection_count > 0)
        + chip("⚠ 주의", f"{caution_count}개", STATUS_META["주의"]["border"])
        + chip("✓ 정상", f"{normal_count}개", STATUS_META["정상"]["border"])
        + chip("전체", f"{total_ct}개")
        + "</div>"
    )
    st.markdown(chips_html, unsafe_allow_html=True)

    if not check_list_df.empty:
        worst = check_list_df.iloc[0]
        worst_ct = worst["컨테이너"]
        worst_issue = worst["이슈"]
    else:
        worst_ct = "-"
        worst_issue = "점검 대상 없음"
    st.caption(
        f"최우선 점검: {worst_ct} / 대표 이슈: {worst_issue} / "
        f"이상값 {abnormal_count:,}건 / 분석 제외 {excluded_count:,}건"
    )


def render_status_cards(container_status_df):
    st.markdown("#### 🧊 컨테이너 상태 카드")
    if container_status_df.empty:
        st.info("컨테이너 상태 데이터가 없습니다.")
        return
    sorted_cards = container_status_df.copy()
    sorted_cards["컨테이너정렬키"] = sorted_cards["컨테이너"].apply(ct_sort_key)
    sorted_cards = sorted_cards.sort_values("컨테이너정렬키").drop(columns=["컨테이너정렬키"])
    cards_per_row = min(5, max(1, len(sorted_cards)))
    for start_idx in range(0, len(sorted_cards), cards_per_row):
        row_cards = sorted_cards.iloc[start_idx:start_idx + cards_per_row]
        cols = st.columns(cards_per_row)
        for idx, (_, row) in enumerate(row_cards.iterrows()):
            style = status_style(row["종합상태"])
            eff_text = "-"
            if pd.notna(row["최저냉동효율"]) and pd.notna(row["최고냉동효율"]):
                eff_text = f"{row['최저냉동효율']:.1f}~{row['최고냉동효율']:.1f}%"
            retention_text = "-"
            if pd.notna(row["최저유지율"]):
                retention_text = f"{row['최저유지율']:.1f}%"
            today_off = "-"
            if pd.notna(row["오늘이탈수"]):
                today_off = f"{int(row['오늘이탈수'])}/3"
            overcool_html = ""
            if bool(row.get("과냉참고", False)):
                overcool_html = (
                    '<div style="margin-top:5px; color:#93c5fd; '
                    'font-size:12px; font-weight:700;">❄ 과냉 참고 (효율 120% 초과)</div>'
                )
            html = f"""
            <div style="
                border: 1.2px solid {style['border']};
                background: {style['bg']};
                border-radius: 12px;
                padding: 12px 12px 10px 12px;
                min-height: 128px;
                margin-bottom: 10px;
                box-shadow: 0 1px 3px rgba(0,0,0,0.16);
            ">
                <div style="display:flex; justify-content:space-between; align-items:center; margin-bottom:7px;">
                    <div style="font-size:20px; font-weight:850; color:white;">{row['컨테이너']}</div>
                    <div style="
                        background:{style['badge']};
                        color:white;
                        border-radius:999px;
                        padding:3px 10px;
                        font-size:12px;
                        font-weight:800;
                        white-space:nowrap;
                    ">{style['icon']} {row['종합상태']}</div>
                </div>
                <div style="color:{style['text']}; font-size:13px; line-height:1.6;">
                    <b>오늘 이탈</b> {today_off}<br>
                    <b>최근3일 이슈</b> {row['최근3일이슈일']}일<br>
                    <b>효율</b> {eff_text}<br>
                    <b>유지율 최저</b> {retention_text}
                </div>
                {overcool_html}
            </div>
            """
            with cols[idx]:
                st.markdown(html, unsafe_allow_html=True)


def render_v4_summary_dashboard(
    daily_status_df,
    recent_df,
    container_status_df,
    check_list_df,
    abnormal_count,
    excluded_count,
    dashboard_start,
    dashboard_today
):
    st.header("🚦 V4 컨테이너 상태 대시보드")
    st.caption(
        f"상단 대시보드는 한국 날짜 기준 최근 3일({dashboard_start} ~ {dashboard_today})만 사용합니다. "
        "상세 분석 탭은 전체 기간을 유지합니다."
    )
    render_summary_chips(container_status_df, check_list_df, abnormal_count, excluded_count)
    render_status_cards(container_status_df)

    st.markdown("#### 🔎 우선 점검 리스트")
    if check_list_df.empty:
        st.success("✅ 최근 3일 기준 우선 점검 대상이 없습니다.")
    else:
        st.dataframe(
            check_list_df,
            use_container_width=True,
            height=230
        )

    st.markdown("#### 🗓️ 컨테이너 × 일자 히트맵")
    heatmap_metric = st.segmented_control(
        "히트맵 기준",
        ["종합상태", "최저온도", "냉동효율(%)", "-15℃이하유지율(%)"],
        default="종합상태",
        key="heatmap_metric_control",
    )
    if heatmap_metric is None:
        heatmap_metric = "종합상태"
    fig_heatmap = create_container_heatmap(daily_status_df, heatmap_metric)
    st.plotly_chart(
        fig_heatmap,
        use_container_width=True,
        config={
            "displayModeBar": True,
            "responsive": True
        }
    )


def create_container_heatmap(daily_status_df, heatmap_metric):
    df = daily_status_df.copy()
    df["측정일자_dt"] = pd.to_datetime(df["측정일자"])
    df["날짜"] = df["측정일자_dt"].dt.strftime("%m-%d")
    df["컨테이너정렬키"] = df["컨테이너"].apply(ct_sort_key)
    df = df.sort_values(["컨테이너정렬키", "측정일자_dt"])
    containers = sorted(df["컨테이너"].unique(), key=ct_sort_key)
    dates = sorted(df["날짜"].unique())
    if heatmap_metric == "종합상태":
        # 심각도 순: 정상 < 주의 < 위험 < 긴급점검 < 데이터없음
        status_score = {
            "정상": 0,
            "주의": 1,
            "위험": 2,
            "긴급점검": 3,
            "데이터없음": 4,
        }
        z_df = (
            df.pivot_table(index="컨테이너", columns="날짜", values="상태", aggfunc="last")
            .reindex(index=containers, columns=dates)
            .fillna("데이터없음")
        )
        issue_df = (
            df.pivot_table(index="컨테이너", columns="날짜", values="이슈", aggfunc="last")
            .reindex(index=containers, columns=dates)
            .fillna("데이터 없음")
        )
        z = z_df.replace(status_score).astype(float).values
        text = z_df.values
        issues = issue_df.values
        customdata = np.dstack([text, issues])
        fig = go.Figure(
            data=go.Heatmap(
                z=z,
                x=dates,
                y=containers,
                customdata=customdata,
                xgap=2,
                ygap=2,
                colorscale=[
                    [0.00, "#22c55e"],  # 정상
                    [0.19, "#22c55e"],
                    [0.20, "#eab308"],  # 주의
                    [0.39, "#eab308"],
                    [0.40, "#f97316"],  # 위험
                    [0.59, "#f97316"],
                    [0.60, "#dc2626"],  # 긴급점검 (가장 강한 색)
                    [0.79, "#dc2626"],
                    [0.80, "#6b7280"],  # 데이터 없음
                    [1.00, "#6b7280"],
                ],
                zmin=0,
                zmax=4,
                colorbar=dict(
                    title="Status",
                    tickvals=[0, 1, 2, 3, 4],
                    ticktext=["Normal", "Caution", "Risk", "Emergency", "No data"]
                ),
                hovertemplate=(
                    "Container: %{y}<br>"
                    "Date: %{x}<br>"
                    "Status: %{customdata[0]}<br>"
                    "Issue: %{customdata[1]}"
                    "<extra></extra>"
                )
            )
        )
        fig.update_layout(
            title="Container x Date Status Heatmap",
            height=430,
            margin=dict(l=50, r=30, t=60, b=40),
            plot_bgcolor="#000000",
            paper_bgcolor="rgba(0,0,0,0)",
            xaxis=dict(
                type="category",
                categoryorder="array",
                categoryarray=dates
            ),
            yaxis=dict(
                type="category",
                categoryorder="array",
                categoryarray=containers
            )
        )
        return fig
    value_df = (
        df.pivot_table(index="컨테이너", columns="날짜", values=heatmap_metric, aggfunc="last")
        .reindex(index=containers, columns=dates)
    )
    issue_df = (
        df.pivot_table(index="컨테이너", columns="날짜", values="이슈", aggfunc="last")
        .reindex(index=containers, columns=dates)
        .fillna("데이터 없음")
    )
    z = value_df.astype(float).values
    issues = issue_df.values
    customdata = np.dstack([issues])
    title = METRIC_LABEL_EN.get(heatmap_metric, heatmap_metric)
    if heatmap_metric == "최저온도":
        colorscale = [
            [0.0, "#1d4ed8"],
            [0.35, "#60a5fa"],
            [0.55, "#e0f2fe"],
            [0.72, "#facc15"],
            [1.0, "#dc2626"],
        ]
        zmin = -25
        zmax = -10
    elif heatmap_metric == "냉동효율(%)":
        colorscale = [
            [0.0, "#dc2626"],
            [0.35, "#facc15"],
            [0.50, "#22c55e"],
            [0.75, "#22c55e"],
            [1.0, "#2563eb"],
        ]
        zmin = 40
        zmax = 140
    else:
        colorscale = [
            [0.0, "#dc2626"],
            [0.35, "#facc15"],
            [0.60, "#22c55e"],
            [1.0, "#16a34a"],
        ]
        zmin = 0
        zmax = 100
    fig = go.Figure(
        data=go.Heatmap(
            z=z,
            x=dates,
            y=containers,
            customdata=customdata,
            xgap=2,
            ygap=2,
            colorscale=colorscale,
            zmin=zmin,
            zmax=zmax,
            colorbar=dict(title=title),
            hovertemplate=(
                "Container: %{y}<br>"
                "Date: %{x}<br>"
                + title + ": %{z:.1f}<br>"
                "Issue: %{customdata[0]}"
                "<extra></extra>"
            )
        )
    )
    fig.update_layout(
        title=f"Container x Date {title} Heatmap",
        height=430,
        margin=dict(l=50, r=30, t=60, b=40),
        plot_bgcolor="#000000",
        paper_bgcolor="rgba(0,0,0,0)",
        xaxis=dict(
            type="category",
            categoryorder="array",
            categoryarray=dates
        ),
        yaxis=dict(
            type="category",
            categoryorder="array",
            categoryarray=containers
        )
    )
    return fig


# =========================================================
# 날짜별 지표 요약 테이블
# =========================================================
@st.cache_data(show_spinner=False)
def build_metric_table(df_summary):
    metrics = ["최저온도", "냉동효율(%)", "-15℃이하유지율(%)"]
    df_filtered = df_summary[
        ["컨테이너", "측정일자"] + metrics
    ].copy()
    for metric in metrics:
        df_filtered[metric] = df_filtered[metric].round(1)
    df_filtered["측정일자"] = pd.to_datetime(df_filtered["측정일자"])
    df_filtered.sort_values("측정일자", inplace=True)
    result_blocks = []
    for metric in metrics:
        pivot = df_filtered.pivot_table(
            index="측정일자",
            columns="컨테이너",
            values=metric,
            aggfunc="last"
        )
        block = pivot.T.copy()
        block.insert(0, "지표", metric)
        block.insert(0, "컨테이너", block.index)
        result_blocks.append(block)
    df_final = pd.concat(result_blocks, axis=0)
    new_columns = []
    for col in df_final.columns:
        if isinstance(col, pd.Timestamp):
            new_columns.append(col.strftime("%m월 %d일"))
        else:
            new_columns.append(col)
    df_final.columns = new_columns
    df_final["컨테이너정렬키"] = df_final["컨테이너"].apply(ct_sort_key)
    metric_order = {
        "최저온도": 1,
        "냉동효율(%)": 2,
        "-15℃이하유지율(%)": 3
    }
    df_final["지표정렬키"] = df_final["지표"].map(metric_order)
    df_final = (
        df_final
        .sort_values(["컨테이너정렬키", "지표정렬키"])
        .drop(columns=["컨테이너정렬키", "지표정렬키"])
        .reset_index(drop=True)
    )
    date_cols = [col for col in df_final.columns if col not in ["컨테이너", "지표"]]
    for col in date_cols:
        df_final[col] = pd.to_numeric(df_final[col], errors="coerce").round(1)
    return df_final


def style_metric_table(df):
    def style_row(row):
        styles = [""] * len(row)
        metric = row.get("지표", "")
        for idx, col in enumerate(row.index):
            if col in ["컨테이너", "지표"]:
                continue
            value = row[col]
            if pd.isna(value):
                continue
            try:
                value = float(value)
            except Exception:
                continue
            if metric == "최저온도":
                if value >= TH["min_temp_off"]:
                    styles[idx] = "color: #ff4b4b; font-weight: 700;"
            elif metric == "냉동효율(%)":
                if value < TH["eff_low"]:
                    styles[idx] = "color: #ff4b4b; font-weight: 700;"
                elif value > TH["eff_high"]:
                    styles[idx] = "color: #3b82f6; font-weight: 700;"
            elif metric == "-15℃이하유지율(%)":
                if value <= TH["retention_off"]:
                    styles[idx] = "color: #ff4b4b; font-weight: 700;"
        return styles
    return df.style.apply(style_row, axis=1).format(
        formatter="{:.1f}",
        subset=[col for col in df.columns if col not in ["컨테이너", "지표"]]
    )


# =========================================================
# 그래프 생성
# =========================================================
def create_temperature_chart_plotly(plot_df, title):
    """화면용 온도 추이 그래프 (Plotly, 다크 테마 통일)"""
    plot_df = plot_df.sort_values("측정일시")
    fig = go.Figure()
    # 0℃ 미만 면적
    area_y = np.minimum(plot_df["온도"].values, 0)
    fig.add_trace(
        go.Scatter(
            x=plot_df["측정일시"],
            y=area_y,
            mode="none",
            fill="tozeroy",
            fillcolor="rgba(56, 189, 248, 0.22)",
            hoverinfo="skip",
            showlegend=False,
        )
    )
    # 온도 라인
    fig.add_trace(
        go.Scatter(
            x=plot_df["측정일시"],
            y=plot_df["온도"],
            mode="lines",
            name="Temperature",
            line=dict(color="#fb923c", width=1.4),
            hovertemplate="%{x|%m-%d %H:%M}<br>온도: %{y:.1f}℃<extra></extra>",
        )
    )
    # 기준선
    fig.add_hline(y=0, line_dash="dash", line_color="#ef4444", line_width=1,
                  annotation_text="0℃", annotation_font_color="#ef4444")
    fig.add_hline(y=TH["retention_temp"], line_dash="dot", line_color="#22c55e", line_width=1,
                  annotation_text=f"{TH['retention_temp']:.0f}℃", annotation_font_color="#22c55e")
    fig.add_hline(y=-TH["target_temp"], line_dash="dash", line_color="#3b82f6", line_width=1,
                  annotation_text=f"{-TH['target_temp']:.0f}℃", annotation_font_color="#3b82f6")
    fig.update_layout(
        title=title,
        height=380,
        margin=dict(l=40, r=30, t=60, b=40),
        yaxis=dict(title="Temp (°C)", range=[-25, 38], showgrid=True),
        xaxis=dict(title="Timestamp", showgrid=True),
        hovermode="x unified",
        showlegend=False,
    )
    return fig


def create_temperature_chart(plot_df, title, figsize=(6, 2.3)):
    """엑셀 내장용 온도 추이 그래프 (matplotlib 유지)"""
    fig, ax = plt.subplots(figsize=figsize)
    plot_df = plot_df.sort_values("측정일시")
    ax.plot(
        plot_df["측정일시"],
        plot_df["온도"],
        label="Temperature",
        color="orange",
        marker="o",
        markersize=0.5,
        linewidth=0.6
    )
    ax.axhline(0, color="red", linestyle="--", linewidth=0.6, label="0°C")
    ax.axhline(TH["retention_temp"], color="green", linestyle=":", linewidth=0.6, label="-15°C")
    ax.axhline(-TH["target_temp"], color="blue", linestyle="--", linewidth=0.6, label="-18°C")
    ax.fill_between(
        plot_df["측정일시"],
        plot_df["온도"],
        0,
        where=(plot_df["온도"] < 0),
        interpolate=True,
        color="skyblue",
        alpha=0.35,
        label="Area (Temp < 0°C)"
    )
    ax.set_ylim(-22, 36)
    ax.set_title(title, fontsize=7)
    ax.set_xlabel("Timestamp", fontsize=6)
    ax.set_ylabel("Temp (°C)", fontsize=6)
    ax.tick_params(axis="x", labelsize=5, rotation=25)
    ax.tick_params(axis="y", labelsize=5)
    ax.legend(loc="upper right", fontsize=5)
    ax.grid(True, linewidth=0.3, alpha=0.6)
    locator = mdates.AutoDateLocator(minticks=4, maxticks=8)
    formatter = mdates.ConciseDateFormatter(locator)
    ax.xaxis.set_major_locator(locator)
    ax.xaxis.set_major_formatter(formatter)
    fig.tight_layout()
    return fig


def create_daily_metric_compare_chart(df_summary, selected_metric, selected_containers):
    selected_metric_label = METRIC_LABEL_EN.get(selected_metric, selected_metric)
    plot_df = df_summary[
        df_summary["컨테이너"].isin(selected_containers)
    ].copy()
    plot_df["측정일자"] = pd.to_datetime(plot_df["측정일자"])
    plot_df = plot_df.sort_values(["컨테이너", "측정일자"])
    fig = go.Figure()
    for ct in sorted(selected_containers, key=ct_sort_key):
        ct_df = plot_df[plot_df["컨테이너"] == ct].copy()
        if ct_df.empty:
            continue
        fig.add_trace(
            go.Scatter(
                x=ct_df["측정일자"],
                y=ct_df[selected_metric],
                mode="lines+markers",
                name=ct,
                line=dict(width=2),
                marker=dict(size=6),
                hovertemplate=(
                    "<b>Container: " + ct + "</b><br>"
                    "Date: %{x|%Y-%m-%d}<br>"
                    + selected_metric_label + ": %{y:.1f}"
                    "<extra></extra>"
                )
            )
        )
    shapes = []
    if selected_metric == "-15℃이하유지율(%)":
        shapes.append({
            "type": "line",
            "xref": "paper",
            "x0": 0,
            "x1": 1,
            "yref": "y",
            "y0": TH["retention_off"],
            "y1": TH["retention_off"],
            "line": {"dash": "dash", "width": 1}
        })
        y_range = [0, 105]
    elif selected_metric == "냉동효율(%)":
        shapes.extend([
            {
                "type": "line",
                "xref": "paper",
                "x0": 0,
                "x1": 1,
                "yref": "y",
                "y0": TH["eff_low"],
                "y1": TH["eff_low"],
                "line": {"dash": "dash", "width": 1}
            },
            {
                "type": "line",
                "xref": "paper",
                "x0": 0,
                "x1": 1,
                "yref": "y",
                "y0": TH["eff_high"],
                "y1": TH["eff_high"],
                "line": {"dash": "dash", "width": 1}
            }
        ])
        y_range = None
    elif selected_metric == "최저온도":
        shapes.extend([
            {
                "type": "line",
                "xref": "paper",
                "x0": 0,
                "x1": 1,
                "yref": "y",
                "y0": TH["min_temp_off"],
                "y1": TH["min_temp_off"],
                "line": {"dash": "dot", "width": 1}
            },
            {
                "type": "line",
                "xref": "paper",
                "x0": 0,
                "x1": 1,
                "yref": "y",
                "y0": -TH["target_temp"],
                "y1": -TH["target_temp"],
                "line": {"dash": "dash", "width": 1}
            }
        ])
        y_range = None
    else:
        y_range = None
    fig.update_layout(
        title=f"Daily {selected_metric_label} Trend",
        xaxis_title="Date",
        yaxis_title=selected_metric_label,
        height=420,
        margin=dict(l=40, r=30, t=60, b=40),
        hovermode="closest",
        legend=dict(
            orientation="h",
            yanchor="bottom",
            y=1.02,
            xanchor="right",
            x=1,
            font=dict(size=10)
        ),
        shapes=shapes
    )
    fig.update_xaxes(
        tickformat="%m-%d",
        showgrid=True
    )
    fig.update_yaxes(
        showgrid=True
    )
    if y_range is not None:
        fig.update_yaxes(range=y_range)
    return fig


@st.cache_data(show_spinner=False)
def build_selected_metric_pivot_table(df_summary, selected_metric, selected_containers_tuple):
    selected_containers = list(selected_containers_tuple)
    table_df = df_summary[
        df_summary["컨테이너"].isin(selected_containers)
    ].copy()
    table_df["측정일자"] = pd.to_datetime(table_df["측정일자"])
    table_df[selected_metric] = table_df[selected_metric].round(1)
    pivot = table_df.pivot_table(
        index="측정일자",
        columns="컨테이너",
        values=selected_metric,
        aggfunc="last"
    )
    selected_containers_sorted = sorted(selected_containers, key=ct_sort_key)
    pivot = pivot.reindex(columns=selected_containers_sorted)
    pivot = pivot.sort_index()
    pivot.index = pivot.index.strftime("%m월 %d일")
    pivot_reset = pivot.reset_index()
    pivot_reset.rename(columns={"index": "측정일자"}, inplace=True)
    return pivot_reset


# =========================================================
# 엑셀 다운로드 생성 (캐싱: 동일 데이터면 재생성 안 함)
# =========================================================
@st.cache_data(show_spinner=False)
def create_excel_download(
    df_summary,
    df_long,
    df_metric_table,
    df_abnormal,
    container_status_df=None,
    check_list_df=None,
    daily_status_df=None
) -> bytes:
    output = io.BytesIO()
    with pd.ExcelWriter(output, engine="openpyxl") as writer:
        df_summary.to_excel(writer, index=False, sheet_name="Summary")
        workbook = writer.book
        worksheet_summary = writer.sheets["Summary"]
        auto_adjust_excel_column_width(worksheet_summary)
        graph_ws = workbook.create_sheet(title="chart")
        row_offset = 2
        ct_list = sorted(df_long["컨테이너"].unique(), key=ct_sort_key)
        for ct in ct_list:
            plot_df = df_long[df_long["컨테이너"] == ct].copy()
            if plot_df.empty:
                continue
            fig = create_temperature_chart(
                plot_df=plot_df,
                title=f"{ct} Temperature Profile",
                figsize=(8, 3)
            )
            img_buf = io.BytesIO()
            fig.savefig(img_buf, format="png", bbox_inches="tight", dpi=150)
            plt.close(fig)
            img_buf.seek(0)
            img = XLImage(img_buf)
            img.anchor = f"B{row_offset}"
            graph_ws.add_image(img)
            row_offset += 18
        df_metric_table.to_excel(writer, index=False, sheet_name="table")
        worksheet_table = writer.sheets["table"]
        auto_adjust_excel_column_width(worksheet_table)
        if df_abnormal.empty:
            df_abnormal_export = pd.DataFrame({
                "메시지": ["이상값 없음"]
            })
        else:
            df_abnormal_export = df_abnormal.copy()
        df_abnormal_export.to_excel(writer, index=False, sheet_name="abnormal_values")
        worksheet_abnormal = writer.sheets["abnormal_values"]
        auto_adjust_excel_column_width(worksheet_abnormal)
        if container_status_df is not None:
            container_status_df.to_excel(writer, index=False, sheet_name="v4_container_status")
            auto_adjust_excel_column_width(writer.sheets["v4_container_status"])
        if check_list_df is not None:
            if check_list_df.empty:
                pd.DataFrame({"메시지": ["점검 대상 없음"]}).to_excel(
                    writer,
                    index=False,
                    sheet_name="v4_check_list"
                )
            else:
                check_list_df.to_excel(writer, index=False, sheet_name="v4_check_list")
            auto_adjust_excel_column_width(writer.sheets["v4_check_list"])
        if daily_status_df is not None:
            export_daily = daily_status_df.copy()
            if "측정일자_dt" in export_daily.columns:
                export_daily = export_daily.drop(columns=["측정일자_dt"])
            export_daily.to_excel(writer, index=False, sheet_name="v4_daily_status")
            auto_adjust_excel_column_width(writer.sheets["v4_daily_status"])
    output.seek(0)
    return output.getvalue()


# =========================================================
# Streamlit 실행부 (분석 프로그램)
# =========================================================
uploaded_file = st.file_uploader(
    "신규 플랫폼 CSV 파일을 업로드하세요",
    type="csv"
)
if uploaded_file is not None:
    try:
        file_bytes = uploaded_file.getvalue()
        with st.spinner("CSV 데이터를 분석 중입니다..."):
            df_raw, df_long, df_abnormal, abnormal_count, excluded_count = (
                preprocess_new_platform_csv(file_bytes)
            )
            if df_long.empty:
                st.error("분석 가능한 온도 데이터가 없습니다. CSV 파일 또는 이상값 기준을 확인해 주세요.")
                st.stop()
            df_summary = calculate_summary(df_long)
            if df_summary.empty:
                st.error("요약 결과를 생성하지 못했습니다.")
                st.stop()
            df_metric_table = build_metric_table(df_summary)
            today = get_today_kst()
            daily_status_df, recent_df, container_status_df, check_list_df, dashboard_start, dashboard_today = (
                build_status_tables(df_summary, today)
            )
        period_dates = pd.date_range(
            start=dashboard_start,
            end=dashboard_today,
            freq="D"
        ).date.tolist()

        report_meta = report_payload.build_report_meta(
            report_date=dashboard_today,
            period_start=dashboard_start,
            period_end=dashboard_today,
            period_dates=period_dates,
            uploaded_filename=uploaded_file.name,
            dashboard_period_days=len(period_dates),
        )

        summary_payload = report_payload.build_summary(container_status_df)
        heatmap_rows_payload = report_payload.build_heatmap_rows(
            daily_status_df=daily_status_df,
            container_status_df=container_status_df,
            period_dates=period_dates,
        )
        check_list_payload = report_payload.build_check_list(
            check_list_df=check_list_df,
            container_status_df=container_status_df,
        )
        metrics_payload = report_payload.build_metrics(container_status_df)
        data_quality_payload = report_payload.build_data_quality(
            df_raw=df_raw,
            df_long=df_long,
            abnormal_count=abnormal_count,
            excluded_count=excluded_count,
        )
        daily_summary_rows_payload = report_payload.build_daily_summary_rows(
            df_summary=df_summary,
        )

        report_payload_json = report_payload.build_report_payload(
            mode="draft",
            report_meta=report_meta,
            summary=summary_payload,
            heatmap_rows=heatmap_rows_payload,
            check_list=check_list_payload,
            metrics=metrics_payload,
            data_quality=data_quality_payload,
            daily_summary_rows=daily_summary_rows_payload,
        )
        current_payload_key = "|".join([
            str(report_meta.get("report_date", "")),
            str(report_meta.get("period_start", "")),
            str(report_meta.get("period_end", "")),
            str(report_meta.get("uploaded_filename", "")),
            str(len(daily_summary_rows_payload)),
            str(len(heatmap_rows_payload)),
            str(len(check_list_payload)),
            str(len(metrics_payload)),
            str(data_quality_payload.get("valid_count", 0)),
            str(data_quality_payload.get("invalid_count", 0)),
        ])

        if st.session_state.get("report_payload_key") != current_payload_key:
            st.session_state["main_preview_result"] = None
            st.session_state["main_preview_ok"] = False
            st.session_state["main_preview_attempted"] = False
            st.session_state["send_confirmed"] = False
            st.session_state["send_result"] = None
            st.session_state["send_ok"] = False

        st.session_state["report_payload_key"] = current_payload_key

        # 자동보고서 관리 화면에서 재사용할 분석 결과 저장
        st.session_state["analysis_done"] = True
        st.session_state["report_payload_json"] = report_payload_json
        st.session_state["report_meta"] = report_meta
        st.session_state["summary_payload"] = summary_payload
        st.session_state["heatmap_rows_payload"] = heatmap_rows_payload
        st.session_state["check_list_payload"] = check_list_payload
        st.session_state["metrics_payload"] = metrics_payload
        st.session_state["data_quality_payload"] = data_quality_payload
        st.session_state["daily_summary_rows_payload"] = daily_summary_rows_payload

        st.success("✅ 신규 플랫폼 CSV 데이터 불러오기 및 전처리 완료")

        # =========================================================
        # 탭 구조: 대시보드 / 상세 분석 / 그래프
        # =========================================================
        tab_dashboard, tab_detail, tab_chart = st.tabs(
            ["🚦 상태 대시보드", "📋 상세 분석", "📈 그래프 분석"]
        )

        # -------------------------------------------------
        # 탭 1: 상태 대시보드
        # -------------------------------------------------
        with tab_dashboard:
            render_v4_summary_dashboard(
                daily_status_df=daily_status_df,
                recent_df=recent_df,
                container_status_df=container_status_df,
                check_list_df=check_list_df,
                abnormal_count=abnormal_count,
                excluded_count=excluded_count,
                dashboard_start=dashboard_start,
                dashboard_today=dashboard_today
            )

        # -------------------------------------------------
        # 탭 2: 상세 분석
        # -------------------------------------------------
        with tab_detail:
            st.header("📋 상세 분석")
            st.caption("전체 기간 기준 상세 분석 결과입니다.")
            min_date = df_long["측정일시"].min()
            max_date = df_long["측정일시"].max()
            ct_count = df_long["컨테이너"].nunique()
            valid_count = len(df_long)
            col_a, col_b, col_c, col_d = st.columns(4)
            with col_a:
                st.metric("측정 시작", min_date.strftime("%Y-%m-%d %H:%M"))
            with col_b:
                st.metric("측정 종료", max_date.strftime("%Y-%m-%d %H:%M"))
            with col_c:
                st.metric("컨테이너 수", f"{ct_count}개")
            with col_d:
                st.metric("유효 측정건수", f"{valid_count:,}건")
            st.metric("이상값 대체 건수", f"{abnormal_count:,}건")
            if abnormal_count > 0:
                with st.expander(
                    f"⚠️ 이상값 목록 확인 ({TH['abnormal_low']:.0f}℃ 미만 또는 {TH['abnormal_high']:.0f}℃ 초과)"
                ):
                    st.dataframe(
                        df_abnormal,
                        use_container_width=True,
                        height=220
                    )

            st.subheader("📊 컨테이너별 날짜별 분석결과")
            st.dataframe(
                df_summary,
                use_container_width=True,
                height=350
            )

            st.subheader("📊 컨테이너별 최저온도 / 냉동효율 / -15℃ 이하 유지율 요약 테이블")
            st.caption(
                f"표시 기준: 최저온도 {TH['min_temp_off']:.0f}℃ 이상 빨간색 / "
                f"냉동효율 {TH['eff_low']:.0f}% 미만 빨간색, {TH['eff_high']:.0f}% 초과 파란색(과냉 참고) / "
                f"-15℃ 이하 유지율 {TH['retention_off']:.0f}% 이하 빨간색"
            )
            styled_metric_table = style_metric_table(df_metric_table)
            st.dataframe(
                styled_metric_table,
                use_container_width=True,
                height=320
            )

            excel_bytes = create_excel_download(
                df_summary=df_summary,
                df_long=df_long,
                df_metric_table=df_metric_table,
                df_abnormal=df_abnormal,
                container_status_df=container_status_df,
                check_list_df=check_list_df,
                daily_status_df=daily_status_df
            )
            start_date = df_long["측정일자"].min().strftime("%y%m%d")
            end_date = df_long["측정일자"].max().strftime("%y%m%d")
            filename = f"{start_date}_{end_date}_freezer_ct_dashboard_v4.xlsx"
            st.download_button(
                label="📥 분석결과 엑셀 다운로드 (V4 대시보드 + 상세 분석 포함)",
                data=excel_bytes,
                file_name=filename,
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
            )

        # -------------------------------------------------
        # 탭 3: 그래프 분석
        # -------------------------------------------------
        with tab_chart:
            st.header("📈 그래프 분석")

            st.subheader("온도 추이 그래프")
            ct_list = sorted(df_summary["컨테이너"].unique(), key=ct_sort_key)
            col1, col2 = st.columns([1, 5])
            with col1:
                selected_container = st.selectbox(
                    "Select Container",
                    ct_list
                )
                available_dates = (
                    df_summary[df_summary["컨테이너"] == selected_container]["측정일자"]
                    .sort_values()
                    .unique()
                    .tolist()
                )
                available_dates.insert(0, "전체")
                selected_date = st.selectbox(
                    "Select Date",
                    available_dates
                )
            if selected_date == "전체":
                plot_df = df_long[
                    df_long["컨테이너"] == selected_container
                ].copy()
                title_suffix = "All Dates"
            else:
                plot_df = df_long[
                    (df_long["컨테이너"] == selected_container)
                    & (df_long["측정일자"] == selected_date)
                ].copy()
                title_suffix = str(selected_date)
            if plot_df.empty:
                st.warning("선택한 조건에 해당하는 그래프 데이터가 없습니다.")
            else:
                fig_temp = create_temperature_chart_plotly(
                    plot_df=plot_df,
                    title=f"{selected_container} | {title_suffix} Temperature Profile",
                )
                with col2:
                    st.plotly_chart(
                        fig_temp,
                        use_container_width=True,
                        config={"displayModeBar": True, "responsive": True}
                    )

            st.divider()
            st.subheader("일자별 지표 비교 그래프")
            st.caption("커서를 그래프의 선 또는 점에 올리면 해당 컨테이너 정보만 표시됩니다.")
            metric_options = [
                "최저온도",
                "냉동효율(%)",
                "-15℃이하유지율(%)"
            ]
            selected_metric = st.segmented_control(
                "지표 선택",
                metric_options,
                default="냉동효율(%)",
                key="daily_metric_control",
            )
            if selected_metric is None:
                selected_metric = "냉동효율(%)"

            selected_containers = st.pills(
                "컨테이너 선택",
                options=ct_list,
                selection_mode="multi",
                default=ct_list,
            )
            st.caption(
                f"Selected containers: {len(selected_containers)} / {len(ct_list)}"
            )

            if not selected_containers:
                st.warning("그래프를 표시하려면 컨테이너를 1개 이상 선택해 주세요.")
            else:
                fig_daily_metric = create_daily_metric_compare_chart(
                    df_summary=df_summary,
                    selected_metric=selected_metric,
                    selected_containers=selected_containers
                )
                st.plotly_chart(
                    fig_daily_metric,
                    use_container_width=True,
                    config={
                        "displayModeBar": True,
                        "responsive": True
                    }
                )
                selected_metric_table = build_selected_metric_pivot_table(
                    df_summary=df_summary,
                    selected_metric=selected_metric,
                    selected_containers_tuple=tuple(selected_containers)
                )
                st.markdown(f"#### 선택 지표 테이블: {selected_metric}")
                st.dataframe(
                    selected_metric_table,
                    use_container_width=True,
                    height=260
                )
    except Exception as e:
        st.error("CSV 처리 중 오류가 발생했습니다.")
        st.exception(e)
